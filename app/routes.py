# app/routes.py
import shutil
import csv
import io
from flask import Response
# ... el resto de tus importaciones ...
from flask import Blueprint, render_template, url_for, flash, redirect, request, jsonify, current_app, send_from_directory
from flask_login import login_user, current_user, logout_user, login_required
from functools import wraps
from wtforms.validators import ValidationError
from app import db, bcrypt
from app.models import Usuario, Persona, Dpto, TipoControl, TipoPersona, Registro, Setting
from app.forms import (LoginForm, DptoForm, TipoControlForm, TipoPersonaForm, PersonaForm, UsuarioForm)
from app.utils import generate_qr_code
import os
import secrets
from PIL import Image
from datetime import date, datetime
import base64
from io import BytesIO
from werkzeug.utils import secure_filename
import subprocess
from .utils import convert_utc_to_local

from openpyxl import load_workbook


bp = Blueprint('main', __name__)

# --- Decoradores de roles ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol.nombre != 'Administrador':
            flash('Acceso no autorizado. Se requiere rol de Administrador.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

# --- Rutas de Autenticación y Principales ---
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('main.index'))
    form = LoginForm()
    if form.validate_on_submit():
        user = Usuario.query.filter_by(username=form.username.data).first()
        if user and bcrypt.check_password_hash(user.password, form.password.data):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('main.index'))
        else:
            flash('Login fallido. Por favor, verifica tu usuario y contraseña.', 'danger')
    return render_template('login.html', title='Login', form=form)

@bp.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('main.login'))

@bp.route('/')
@login_required
def index():
    return redirect(url_for('main.registro_almuerzo'))

# --- Ruta principal de Registro de Almuerzos ---
@bp.route('/registro')
@login_required
def registro_almuerzo():
    # Obtener los últimos 10 registros del día
    today = date.today()
    registros_hoy = Registro.query.filter(db.func.date(Registro.fecha_hora_registro) == today)\
                                  .order_by(Registro.fecha_hora_registro.desc()).limit(10).all()
    return render_template('registro_almuerzo.html', title='Registro de Almuerzo', registros=registros_hoy)

# app/routes.py

@bp.route('/procesar_registro', methods=['POST'])
@login_required
def procesar_registro():
    id_persona = request.form.get('id_persona')
    if not id_persona:
        return jsonify({'success': False, 'message': 'Debe ingresar un código.'})

    persona = Persona.query.get(id_persona)
    if not persona:
        return jsonify({'success': False, 'message': 'Código no encontrado.'})

    if persona.control_ref and persona.control_ref.nombre_control.lower() == 'no aplica':
        mensaje = f'"{persona.nombre_persona}" no está habilitado para tomar almuerzo.'
        return jsonify({'success': False, 'message': mensaje})

    hoy_inicio = datetime.combine(date.today(), datetime.min.time())
    registro_existente = Registro.query.filter(Registro.persona_id == id_persona, Registro.fecha_hora_registro >= hoy_inicio).first()
    if registro_existente:
        return jsonify({'success': False, 'message': '¡Ya tomó almuerzo hoy!'})

    nuevo_registro = Registro(
        persona_id=persona.id_persona,
        tipo_control_id=persona.control_id
    )
    db.session.add(nuevo_registro)
    db.session.commit()

    # === INICIO DE LA LÓGICA CORREGIDA Y MÁS ROBUSTA ===
    imprime_tickets_enabled = False # Por defecto, no imprimimos (más seguro)
    imprime_setting = Setting.query.get('IMPRIME_TICKETS')
    
    if imprime_setting:
        # Comparamos el valor de la BD de forma insensible a mayúsculas/minúsculas
        if imprime_setting.value.lower() == 'true':
            imprime_tickets_enabled = True
    # Si 'imprime_setting' no existe, se mantiene en False.
    # === FIN DE LA LÓGICA CORREGIDA ===
    # === CAMBIO CLAVE ===
    # Convertir la hora del nuevo registro a la hora local ANTES de enviarla al frontend
    hora_local_registro = convert_utc_to_local(nuevo_registro.fecha_hora_registro)
    registro_data = {
        'id_registro': nuevo_registro.id_registro,
        'id_persona': persona.id_persona,
        'nombre_persona': persona.nombre_persona,
        'dpto': persona.dpto_ref.nombre_dpto,
        'tipo_control': persona.control_ref.nombre_control,
         # Usar la hora local formateada
        'fecha_hora': hora_local_registro.strftime('%d/%m/%Y %I:%M:%S %p'),
        'imprime_ticket': imprime_tickets_enabled
    }

    return jsonify({'success': True, 'message': 'Registro exitoso.', 'registro': registro_data})

@bp.route('/imprimir_ticket/<int:id_registro>')
@login_required
def imprimir_ticket(id_registro):
    registro = Registro.query.get_or_404(id_registro)
    persona = registro.persona_ref
    
    # === CAMBIO CLAVE ===
    # Convertir la hora del registro a hora local
    hora_local_registro = convert_utc_to_local(registro.fecha_hora_registro)
    fecha_hora_str = hora_local_registro.strftime('%d/%m/%Y %I:%M:%S %p')
    
    ticket_data_str = (
        f"ID: {persona.id_persona}\n"
        f"Nombre: {persona.nombre_persona}\n"
        f"Dpto: {persona.dpto_ref.nombre_dpto}\n"
        f"Control: {registro.tipo_control_ref.nombre_control}\n"
        f"Fecha: {fecha_hora_str}"
    )
    
    qr_code_b64 = generate_qr_code(ticket_data_str)
    # === CAMBIO CLAVE: Leer el nombre del colegio desde la BD ===
    nombre_colegio_setting = Setting.query.get('NOMBRE_COLEGIO')
    nombre_colegio = nombre_colegio_setting.value if nombre_colegio_setting else "Colegio Privado"
    # =========================================================

    return render_template('ticket.html',
                           registro=registro,
                           persona=persona,
                           fecha_hora=fecha_hora_str,
                           qr_code=qr_code_b64,
                           nombre_colegio=nombre_colegio)

# --- Rutas de Administración (CRUDs) ---

# CRUD Genérico para modelos simples (VERSIÓN CORREGIDA)
def crud_view_factory(model, form_class, template_name, list_title, form_title_singular):
    
    # Ruta para mostrar la lista y el formulario (GET)
    @bp.route(f'/{model.__tablename__}', methods=['GET'])
    @login_required
    @admin_required
    def list_items():
        # Pasamos una instancia del formulario a la plantilla para que pueda renderizar el modal
        # form_instance_name = f"{model.__name__.lower()}_form"
        # === CAMBIO CLAVE: Siempre usaremos la variable 'form' en la plantilla ===
        items = model.query.all()
        template_context = {
            'title': list_title,
            'items': items,
            'form_title': form_title_singular,
            'form': form_class() # Pasamos el formulario con el nombre genérico 'form'
        }
        return render_template(f'admin/{template_name}', **template_context)

    # Ruta dedicada para guardar (crear/editar) (POST)
    @bp.route(f'/{model.__tablename__}/save', methods=['POST'])
    @login_required
    @admin_required
    def save_item():
        # === CAMBIO CLAVE 1: Poblar el formulario con los datos de la petición ===
        form = form_class(request.form)
        
        # Obtenemos el ID del campo oculto del formulario
        # Asumimos que el primer campo definido en la clase del formulario es el ID (HiddenField)
        id_field_name = next(iter(form._fields)) 
        item_id = request.form.get(id_field_name)

        if form.validate_on_submit():
            if item_id:  # Edición (si el ID tiene un valor)
                item = model.query.get(item_id)
                if item:
                    # === CAMBIO CLAVE 2: Usar populate_obj para actualizar el objeto de forma genérica ===
                    form.populate_obj(item)
                    flash(f'{form_title_singular} actualizado correctamente!', 'success')
                else:
                    flash('Error: No se encontró el item para actualizar.', 'danger')
            else:  # Creación (si el ID está vacío)
                item = model()
                # === CAMBIO CLAVE 2: Usar populate_obj para poblar el nuevo objeto ===
                form.populate_obj(item)
                # === LA SOLUCIÓN ESTÁ AQUÍ ===
                # Obtenemos el nombre del campo ID del formulario (ej: 'id_dpto')
                id_field_name = next(iter(form._fields)) 
                # Forzamos el atributo del ID a None en el objeto del modelo.
                # Esto le indica a SQLAlchemy que use el auto-incremento de la BD.
                setattr(item, id_field_name, None)
                # ===========================
                db.session.add(item)
                flash(f'{form_title_singular} creado correctamente!', 'success')
            
            db.session.commit()
        else:
            # Si la validación falla, mostrar los errores
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"Error en el campo '{getattr(form, field).label.text}': {error}", 'danger')

        return redirect(url_for(f'main.list_{model.__tablename__}'))

    # Ruta para eliminar (POST)
    @bp.route(f'/{model.__tablename__}/delete/<int:item_id>', methods=['POST'])
    @login_required
    @admin_required
    def delete_item(item_id):
        item = model.query.get_or_404(item_id)
        db.session.delete(item)
        db.session.commit()
        flash(f'{form_title_singular} eliminado.', 'success')
        return redirect(url_for(f'main.list_{model.__tablename__}'))

    # Renombrar las funciones para que Flask no tenga conflictos
    list_items.__name__ = f'list_{model.__tablename__}'
    save_item.__name__ = f'save_{model.__tablename__}'
    delete_item.__name__ = f'delete_{model.__tablename__}'
    
    return list_items, save_item, delete_item

# Generar las vistas CRUD
# (Esta parte no necesita cambios)
dptos_list, dptos_save, dptos_delete = crud_view_factory(Dpto, DptoForm, 'dptos.html', 'Departamentos', 'Departamento')
tipos_control_list, tipos_control_save, tipos_control_delete = crud_view_factory(TipoControl, TipoControlForm, 'tipos_control.html', 'Tipos de Control', 'Tipo de Control')
tipos_persona_list, tipos_persona_save, tipos_persona_delete = crud_view_factory(TipoPersona, TipoPersonaForm, 'tipos_persona.html', 'Tipos de Persona', 'Tipo de Persona')

# --- CRUD para Personas (es más complejo) ---
def save_picture(form_picture, foto_webcam_data):
    if foto_webcam_data:
        # Decodificar imagen de base64
        header, encoded = foto_webcam_data.split(",", 1)
        image_data = base64.b64decode(encoded)
        image_stream = BytesIO(image_data)
        img = Image.open(image_stream)
        
        random_hex = secrets.token_hex(8)
        _, f_ext = os.path.splitext('webcam.png') # Asumimos png
        picture_fn = random_hex + f_ext
        picture_path = os.path.join(current_app.root_path, current_app.config['UPLOAD_FOLDER'], picture_fn)

        output_size = (250, 250)
        img.thumbnail(output_size)
        img.save(picture_path)

        return picture_fn

    elif form_picture:
        random_hex = secrets.token_hex(8)
        _, f_ext = os.path.splitext(form_picture.filename)
        picture_fn = random_hex + f_ext
        picture_path = os.path.join(current_app.root_path, current_app.config['UPLOAD_FOLDER'], picture_fn)

        output_size = (250, 250)
        i = Image.open(form_picture)
        i.thumbnail(output_size)
        i.save(picture_path)

        return picture_fn
    return None

@bp.route('/personas')
@login_required
@admin_required
def list_personas():
    personas = Persona.query.all()
    return render_template('admin/personas.html', title='Personas', personas=personas)

# app/routes.py

# ... (otras rutas) ...

@bp.route('/personas/nueva', methods=['GET', 'POST'])
@login_required
@admin_required
def nueva_persona():
    form = PersonaForm()
    # Eliminar el validador de unicidad que pusimos en el formulario
    # para manejarlo explícitamente aquí.
    del form.id_persona.validators[-1]
    
    if form.validate_on_submit():
        # Validación de unicidad manual
        if Persona.query.get(form.id_persona.data):
            flash('Ese código de persona ya está registrado. Por favor, elige otro.', 'danger')
            return render_template('admin/form_persona.html', title='Nueva Persona', form=form, legend='Nueva Persona')

        # Procesar y guardar la foto
        picture_file = save_picture(form.foto.data, form.foto_webcam.data)

        # Crear la nueva instancia de Persona
        nueva_persona = Persona(
            id_persona=form.id_persona.data,
            nombre_persona=form.nombre_persona.data,
            sexo=form.sexo.data,
            # Obtener los IDs de los objetos seleccionados en los QuerySelectFields
            dpto_id=form.dpto.data.id_dpto,
            tipo_persona_id=form.tipo_persona.data.id_tipopersona,
            control_id=form.control.data.id_control
        )
        if picture_file:
            nueva_persona.foto = picture_file
        
        db.session.add(nueva_persona)
        db.session.commit()
        
        flash('Persona creada exitosamente!', 'success')
        return redirect(url_for('main.list_personas'))
    
    return render_template('admin/form_persona.html', title='Nueva Persona', form=form, legend='Nueva Persona')


@bp.route('/personas/editar/<id_persona>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_persona(id_persona):
    persona = Persona.query.get_or_404(id_persona)
    # Al crear el formulario, WTForms maneja los datos iniciales,
    # pero para QuerySelectField, es mejor establecer el 'data' antes de validar.
    form = PersonaForm(obj=persona)
    
    if request.method == 'GET':
        # Pre-poblar los campos de selección con los objetos correctos
        form.dpto.data = persona.dpto_ref
        form.tipo_persona.data = persona.tipo_persona_ref
        form.control.data = persona.control_ref
    
    # Quitar el validador de unicidad para la edición
    del form.id_persona.validators[-1]

    if form.validate_on_submit():
        # Actualizar la foto si se proporcionó una nueva
        if form.foto.data or form.foto_webcam.data:
            picture_file = save_picture(form.foto.data, form.foto_webcam.data)
            persona.foto = picture_file
        
        # === SOLUCIÓN CLAVE: Actualizar el objeto 'persona' existente ===
        # En lugar de crear uno nuevo, modificamos el que cargamos de la sesión.
        persona.nombre_persona = form.nombre_persona.data
        persona.sexo = form.sexo.data
        persona.dpto_id = form.dpto.data.id_dpto
        persona.tipo_persona_id = form.tipo_persona.data.id_tipopersona
        persona.control_id = form.control.data.id_control
        
        db.session.commit()
        flash('Persona actualizada exitosamente!', 'success')
        return redirect(url_for('main.list_personas'))
    
    # Hacer el campo de ID de solo lectura en la plantilla
    form.id_persona.render_kw = {'readonly': True}
    return render_template('admin/form_persona.html', title='Editar Persona', form=form, legend=f'Editar Persona: {persona.nombre_persona}')


@bp.route('/personas/delete/<id_persona>', methods=['POST'])
@login_required
@admin_required
def delete_persona(id_persona):
    persona = Persona.query.get_or_404(id_persona)
    # === LA SOLUCIÓN ESTÁ AQUÍ ===
    # Contamos cuántos registros tiene esta persona.
    # El método .count() es muy eficiente para esto.
    if persona.registros: # SQLAlchemy evalúa esto eficientemente
        flash(f'No se puede eliminar a "{persona.nombre_persona}" porque tiene registros de almuerzo asociados.', 'danger')
    else:
        # Si no tiene registros, procedemos con la eliminación.
        db.session.delete(persona)
        db.session.commit()
        flash(f'Persona "{persona.nombre_persona}" eliminada correctamente.', 'success')
    # ============================
    return redirect(url_for('main.list_personas'))
    
# --- CRUD de Usuarios ---
@bp.route('/usuarios')
@login_required
@admin_required
def list_usuarios():
    users = Usuario.query.all()
    form = UsuarioForm()
    return render_template('admin/usuarios.html', title='Usuarios', users=users, form=form)

# app/routes.py

@bp.route('/usuarios/save', methods=['POST'])
@login_required
@admin_required
def save_usuario():
    form = UsuarioForm()
    if form.validate_on_submit():
        user_id = form.id.data
        if user_id: # Editando
            user = Usuario.query.get(user_id)
            user.username = form.username.data
            user.rol_id = form.rol.data.id
            if form.password.data:
                user.password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        else: # Creando
            hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            user = Usuario(username=form.username.data, password=hashed_password, rol_id=form.rol.data.id)
            db.session.add(user)
        db.session.commit()
        flash('Usuario guardado exitosamente.', 'success')
    else:
        # Si hay errores de validación, los mostramos
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error en el campo '{getattr(form, field).label.text}': {error}", 'danger')
    return redirect(url_for('main.list_usuarios'))

@bp.route('/usuarios/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_usuario(user_id):
    user = Usuario.query.get_or_404(user_id)
    # No permitir eliminar al único administrador o a sí mismo si es admin
    if user.rol.nombre == 'Administrador' and Usuario.query.filter_by(rol_id=user.rol_id).count() == 1:
        flash('No se puede eliminar al último administrador.', 'danger')
    elif user.id == current_user.id:
        flash('No puedes eliminarte a ti mismo.', 'danger')
    else:
        db.session.delete(user)
        db.session.commit()
        flash('Usuario eliminado.', 'success')
    return redirect(url_for('main.list_usuarios'))
    
# --- Reportes ---
@bp.route('/reportes', methods=['GET', 'POST'])
@login_required
def reportes():
    if request.method == 'POST':
        tipo_reporte = request.form.get('tipo_reporte')
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        filtro_id = request.form.get('filtro_id')
        
        # Construir la consulta base
        query = Registro.query
        
        # Filtros de fecha
        if fecha_inicio:
            query = query.filter(Registro.fecha_hora_registro >= datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        if fecha_fin:
            # Añadir un día para incluir todo el día de fin
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            query = query.filter(Registro.fecha_hora_registro < fecha_fin_dt.replace(day=fecha_fin_dt.day + 1))
        
        # Filtros específicos
        if tipo_reporte == 'persona' and filtro_id:
            query = query.filter(Registro.persona_id == filtro_id)
        elif tipo_reporte == 'tipo_control' and filtro_id:
            query = query.filter(Registro.tipo_control_id == filtro_id)
        elif tipo_reporte == 'dpto' and filtro_id:
            query = query.join(Persona).filter(Persona.dpto_id == filtro_id)
        elif tipo_reporte == 'tipo_persona' and filtro_id:
            query = query.join(Persona).filter(Persona.tipo_persona_id == filtro_id)
            
        resultados = query.order_by(Registro.fecha_hora_registro.desc()).all()
        
        return render_template('reports/ver_reporte.html',
                               resultados=resultados,
                               tipo_reporte=tipo_reporte,
                               title='Resultados del Reporte')

    # Para el método GET, mostrar el selector
    personas = Persona.query.order_by(Persona.nombre_persona).all()
    tipos_control = TipoControl.query.all()
    dptos = Dpto.query.all()
    tipos_persona = TipoPersona.query.all()
    return render_template('reports/selector_reportes.html',
                           title='Generar Reportes',
                           personas=personas,
                           tipos_control=tipos_control,
                           dptos=dptos,
                           tipos_persona=tipos_persona)

# app/routes.py (al final del archivo)

# --- Rutas de Importación CSV ---

@bp.route('/admin/download_template/<modelo>')
@login_required
@admin_required
def download_template(modelo):
    """Genera y descarga una plantilla CSV para el modelo especificado."""
    
    # Define las cabeceras para cada modelo
    headers = {
        'dptos': ['nombre_dpto'],
        'personas': ['id_persona', 'nombre_persona', 'sexo', 'nombre_dpto', 'nombre_tipopersona', 'nombre_control']
    }

    if modelo not in headers:
        flash('Modelo no válido para la descarga de plantilla.', 'danger')
        return redirect(url_for('main.index'))

    # Usar io.StringIO para crear el archivo CSV en memoria
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Escribir la cabecera
    writer.writerow(headers[modelo])
    
    # Crear la respuesta para el navegador
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename=plantilla_{modelo}.csv"}
    )

# app/routes.py

# app/routes.py

@bp.route('/admin/importar/<modelo>', methods=['GET', 'POST'])
@login_required
@admin_required
def importar_csv(modelo):
    config = {
        'dptos': { 'model': Dpto, 'headers': ['nombre_dpto'], 'list_route': 'main.list_dptos' },
        'personas': { 'model': Persona, 'headers': ['id_persona', 'nombre_persona', 'sexo', 'nombre_dpto', 'nombre_tipopersona', 'nombre_control'], 'list_route': 'main.list_personas' }
    }
    if modelo not in config:
        flash('Modelo no válido para importación.', 'danger')
        return redirect(url_for('main.index'))

    if request.method == 'POST':
        if 'csv_file' not in request.files or not request.files['csv_file'].filename:
            flash('No se seleccionó ningún archivo.', 'warning')
            return redirect(request.url)
        
        file = request.files['csv_file']

        if not file.filename.endswith('.csv'):
            flash('Formato de archivo no válido. Por favor, sube un archivo .csv', 'danger')
            return redirect(request.url)

        try:
            raw_bytes = file.stream.read()
            decoded_string = ""
            try:
                decoded_string = raw_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                decoded_string = raw_bytes.decode('latin-1')
            
            stream = io.StringIO(decoded_string, newline=None)

            # === INICIO DE LA MODIFICACIÓN: LÓGICA CONDICIONAL PARA EL LECTOR CSV ===
            if modelo == 'dptos':
                # Para archivos de una sola columna (dptos), no usamos Sniffer.
                csv_reader = csv.reader(stream)
            else:
                # Para archivos multi-columna (personas), el Sniffer es útil.
                try:
                    dialect = csv.Sniffer().sniff(stream.read(2048)) # Aumentamos un poco el tamaño de la muestra
                    stream.seek(0)
                except csv.Error:
                    # Si el Sniffer falla (ej. archivo vacío o de una columna), asumimos comas.
                    dialect = 'excel'
                csv_reader = csv.reader(stream, dialect)
            # === FIN DE LA MODIFICACIÓN ===
            
            header_from_file = next(csv_reader)
            cleaned_header = [h.strip() for h in header_from_file]

            if cleaned_header != config[modelo]['headers']:
                flash(f"La cabecera del archivo es incorrecta. Debería ser: {', '.join(config[modelo]['headers'])}", 'danger')
                return render_template('admin/importar.html', title=f"Importar {modelo.capitalize()}", modelo=modelo, errors=[])
            
            # --- El resto de tu código de validación y upsert no necesita ningún cambio ---
            errors = []
            created_count = 0
            updated_count = 0
            
            if modelo == 'dptos':
                existing_dptos = {d.nombre_dpto for d in Dpto.query.all()}

            elif modelo == 'personas':
                existing_persons_map = {p.id_persona: p for p in Persona.query.all()}
                dptos_map = {d.nombre_dpto: d.id_dpto for d in Dpto.query.all()}
                tipos_persona_map = {tp.nombre_tipopersona: tp.id_tipopersona for tp in TipoPersona.query.all()}
                tipos_control_map = {tc.nombre_control: tc.id_control for tc in TipoControl.query.all()}
                ids_in_file = set()

            for i, row in enumerate(csv_reader, start=2):
                row_errors = [] 
                
                if modelo == 'dptos':
                    if len(row) != 1 or not row[0]: row_errors.append("Formato incorrecto o nombre de departamento vacío.")
                    elif row[0].strip() in existing_dptos: row_errors.append(f"El departamento '{row[0].strip()}' ya existe.")
                    if not row_errors:
                        db.session.add(Dpto(nombre_dpto=row[0].strip()))
                        existing_dptos.add(row[0].strip())
                        created_count += 1
                elif modelo == 'personas':
                    if len(row) != 6:
                        row_errors.append(f"La fila debe tener 6 columnas, pero tiene {len(row)}.")
                    else:
                        id_p, nombre, sexo, dpto_n, tipo_p_n, control_n = [field.strip() for field in row]
                        
                        if not all([id_p, nombre, sexo, dpto_n, tipo_p_n, control_n]): row_errors.append(f"ID '{id_p}': Todos los campos son obligatorios.")
                        if id_p in ids_in_file: row_errors.append(f"ID '{id_p}': está duplicado dentro del mismo archivo CSV.")
                        if sexo.upper() not in ['M', 'F']: row_errors.append(f"ID '{id_p}': El sexo ('{sexo}') debe ser 'M' o 'F'.")
                        if dpto_n not in dptos_map: row_errors.append(f"ID '{id_p}': El departamento '{dpto_n}' no es válido.")
                        if tipo_p_n not in tipos_persona_map: row_errors.append(f"ID '{id_p}': El tipo de persona '{tipo_p_n}' no es válido.")
                        if control_n not in tipos_control_map: row_errors.append(f"ID '{id_p}': El tipo de control '{control_n}' no es válido.")
                        
                        if not row_errors:
                            ids_in_file.add(id_p)
                            
                            if id_p in existing_persons_map:
                                person_to_update = existing_persons_map[id_p]
                                person_to_update.nombre_persona = nombre
                                person_to_update.sexo = sexo.upper()
                                person_to_update.dpto_id = dptos_map[dpto_n]
                                person_to_update.tipo_persona_id = tipos_persona_map[tipo_p_n]
                                person_to_update.control_id = tipos_control_map[control_n]
                                updated_count += 1
                            else:
                                new_person = Persona(
                                    id_persona=id_p, nombre_persona=nombre, sexo=sexo.upper(),
                                    dpto_id=dptos_map[dpto_n], tipo_persona_id=tipos_persona_map[tipo_p_n], control_id=tipos_control_map[control_n]
                                )
                                db.session.add(new_person)
                                created_count += 1
                
                if row_errors:
                    errors.extend([f"Fila {i}: {e}" for e in row_errors])

            if errors:
                db.session.rollback()
                return render_template('admin/importar.html', title=f"Importar {modelo.capitalize()}", modelo=modelo, errors=errors)
            
            elif created_count == 0 and updated_count == 0:
                flash('El archivo CSV estaba vacío o no contenía datos para crear o actualizar.', 'warning')
                return redirect(request.url)

            else:
                db.session.commit()
                success_message = []
                if created_count > 0: success_message.append(f"{created_count} personas creadas")
                if updated_count > 0: success_message.append(f"{updated_count} personas actualizadas")
                flash(f"Importación exitosa: {', '.join(success_message)}.", 'success')
                return redirect(url_for(config[modelo]['list_route']))

        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado al procesar el archivo: {e}', 'danger')
            return redirect(request.url)

    return render_template('admin/importar.html', title=f"Importar {modelo.capitalize()}", modelo=modelo)

# app/routes.py (añadir al final del archivo)

@bp.route('/registro/delete/<int:id_registro>', methods=['POST'])
@login_required
@admin_required
def delete_registro(id_registro):
    """Elimina un registro de almuerzo específico. Solo para administradores."""
    registro = Registro.query.get_or_404(id_registro)
    
    db.session.delete(registro)
    db.session.commit()
    
    flash(f'Registro de {registro.persona_ref.nombre_persona} a las {registro.fecha_hora_registro.strftime("%H:%M:%S")} ha sido eliminado.', 'success')
    
    # Redirigir al usuario a la página desde la que vino (ya sea el registro o un reporte)
    return redirect(request.referrer or url_for('main.registro_almuerzo'))

# app/routes.py (añadir al final del archivo)
@bp.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def settings():
    if request.method == 'POST':
        # --- Manejo de campos de texto ---
        settings_to_update = {
            'IMPRIME_TICKETS': 'True' if 'imprime_tickets' in request.form else 'False',
            'NOMBRE_COLEGIO': request.form.get('nombre_colegio', "Colegio Privado 'El Saber'")
        }

        for key, value in settings_to_update.items():
            setting = Setting.query.get(key)
            if setting:
                setting.value = value
        
        # --- Manejo de la subida del logo ---
        if 'logo_file' in request.files:
            file = request.files['logo_file']
            if file and file.filename != '':
                # Validar extensión
                allowed_extensions = {'png', 'jpg', 'jpeg', 'svg'}
                if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in allowed_extensions:
                    # Guardar el archivo de forma segura
                    filename = secure_filename(file.filename)
                    logo_path = os.path.join(current_app.root_path, 'static/logos', filename)
                    file.save(logo_path)
                    
                    # Actualizar la entrada en la base de datos
                    logo_setting = Setting.query.get('LOGO_FILENAME')
                    if logo_setting:
                        logo_setting.value = filename
                    
                    flash('Logo actualizado exitosamente.', 'success')
                else:
                    flash('Formato de archivo de logo no permitido. Usa png, jpg, jpeg o svg.', 'danger')

        db.session.commit()
        flash('Configuración guardada exitosamente.', 'success')
        return redirect(url_for('main.settings'))

    # Para peticiones GET (no cambia)
    settings_from_db = Setting.query.all()
    settings_dict = {setting.key: setting.value for setting in settings_from_db}
    
    return render_template('admin/settings.html', title='Configuración General', settings=settings_dict)

# app/routes.py (añadir al final del archivo)

# app/routes.py

# --- Rutas de Backup y Restauración para MySQL (Reconstruido) ---

@bp.route('/admin/backup_restore', methods=['GET'])
@login_required
@admin_required
def backup_restore():
    """Muestra la página de gestión de copias de seguridad."""
    backup_folder = current_app.config['BACKUP_FOLDER']
    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)
    
    # Buscar archivos .sql.gz
    server_backups = sorted(
        [f for f in os.listdir(backup_folder) if f.endswith('.sql.gz')],
        reverse=True
    )
    return render_template('admin/backup_restore.html', title='Backup y Restauración', server_backups=server_backups)


def _run_mysql_command(command):
    """Función auxiliar para ejecutar comandos de shell y manejar errores."""
    db_config = current_app.config.get('DB_CONFIG', {})
    safe_command = command.replace(db_config.get('password', 'PASSWORD'), '********')
    
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    stdout, stderr = process.communicate()

    if process.returncode != 0:
        raise RuntimeError(f"Error ejecutando comando: {safe_command}\nError: {stderr.strip()}")
    return stdout


@bp.route('/admin/backup/server', methods=['POST'])
@login_required
@admin_required
def backup_to_server():
    """Crea un backup de la base de datos MySQL y la guarda en el servidor."""
    db_config = current_app.config.get('DB_CONFIG')
    if not db_config:
        flash('La configuración de la base de datos no es válida para esta operación.', 'danger')
        return redirect(url_for('main.backup_restore'))

    try:
        backup_folder = current_app.config['BACKUP_FOLDER']
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup_filename = f"backup_{timestamp}.sql.gz"
        backup_path = os.path.join(backup_folder, backup_filename)

        # Leer las rutas desde la configuración de la app
        mysqldump_cmd = current_app.config['MYSQLDUMP_PATH']
        gzip_cmd = current_app.config['GZIP_PATH']
        
        command = (
            f"{mysqldump_cmd} --user={db_config['user']} --password='{db_config['password']}' "
            f"--host={db_config['host']} --single-transaction --routines --triggers "
            f"{db_config['database']} | {gzip_cmd} > {backup_path}"
        )
        
        _run_mysql_command(command)
        flash(f'Copia de seguridad "{backup_filename}" creada exitosamente.', 'success')
            
    except Exception as e:
        flash(f'Error al crear la copia de seguridad: {e}', 'danger')
        
    return redirect(url_for('main.backup_restore'))


@bp.route('/admin/backup/download')
@login_required
@admin_required
def backup_download():
    """Crea un backup de MySQL y lo ofrece para descargar."""
    db_config = current_app.config.get('DB_CONFIG')
    if not db_config:
        flash('La configuración de la base de datos no es válida para esta operación.', 'danger')
        return redirect(url_for('main.backup_restore'))

    try:
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        backup_filename = f"backup_{timestamp}.sql.gz"
        
        mysqldump_cmd = current_app.config['MYSQLDUMP_PATH']
        gzip_cmd = current_app.config['GZIP_PATH']

        command = (
            f"{mysqldump_cmd} --user={db_config['user']} --password='{db_config['password']}' "
            f"--host={db_config['host']} --single-transaction --routines --triggers "
            f"{db_config['database']} | {gzip_cmd}"
        )
        
        process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = process.communicate()

        if process.returncode != 0:
            raise RuntimeError(f"Error al generar el backup: {stderr.decode('utf-8')}")

        return Response(
            stdout,
            mimetype="application/gzip",
            headers={"Content-disposition": f"attachment; filename={backup_filename}"}
        )
        
    except Exception as e:
        flash(f'Error inesperado al generar la descarga: {e}', 'danger')
        return redirect(url_for('main.backup_restore'))


@bp.route('/admin/restore/upload', methods=['POST'])
@login_required
@admin_required
def restore_from_upload():
    """Restaura la base de datos MySQL desde un archivo .sql.gz subido."""
    file = request.files.get('backup_file')
    if not file or not file.filename:
        flash('No se seleccionó ningún archivo para restaurar.', 'warning')
        return redirect(url_for('main.backup_restore'))
    
    if not file.filename.endswith('.sql.gz'):
        flash('Archivo no válido. Solo se pueden restaurar archivos .sql.gz', 'danger')
        return redirect(url_for('main.backup_restore'))

    db_config = current_app.config.get('DB_CONFIG')
    if not db_config:
        flash('La configuración de la base de datos no es válida para esta operación.', 'danger')
        return redirect(url_for('main.backup_restore'))

    try:
        gunzip_cmd = current_app.config['GUNZIP_PATH']
        mysql_cmd = current_app.config['MYSQL_PATH']

        command = (
            f"{gunzip_cmd} | {mysql_cmd} --user={db_config['user']} "
            f"--password='{db_config['password']}' --host={db_config['host']} {db_config['database']}"
        )
        
        process = subprocess.Popen(command, shell=True, stdin=file.stream, stderr=subprocess.PIPE, text=True)
        _, stderr = process.communicate()

        if process.returncode != 0:
            raise RuntimeError(f"Error al restaurar: {stderr}")

        flash('Restauración desde archivo local completada.', 'success')
        flash('IMPORTANTE: Puede ser necesario reiniciar la aplicación para que todos los cambios surtan efecto.', 'warning')
            
    except Exception as e:
        flash(f'Error inesperado al restaurar: {e}', 'danger')

    return redirect(url_for('main.backup_restore'))


@bp.route('/admin/restore/server', methods=['POST'])
@login_required
@admin_required
def restore_from_server():
    """Restaura la base de datos MySQL desde un archivo del servidor."""
    backup_filename = request.form.get('backup_filename')
    if not backup_filename:
        flash('No se seleccionó ningún archivo de backup del servidor.', 'warning')
        return redirect(url_for('main.backup_restore'))
    
    db_config = current_app.config.get('DB_CONFIG')
    if not db_config:
        flash('La configuración de la base de datos no es válida para esta operación.', 'danger')
        return redirect(url_for('main.backup_restore'))

    try:
        backup_folder = current_app.config['BACKUP_FOLDER']
        backup_path = os.path.join(backup_folder, secure_filename(backup_filename))
        
        if not os.path.exists(backup_path):
            raise FileNotFoundError('El archivo de backup seleccionado no existe.')

        gunzip_cmd = current_app.config['GUNZIP_PATH']
        mysql_cmd = current_app.config['MYSQL_PATH']

        command = (
            f"{gunzip_cmd} < {backup_path} | {mysql_cmd} --user={db_config['user']} "
            f"--password='{db_config['password']}' --host={db_config['host']} {db_config['database']}"
        )
        
        _run_mysql_command(command)
        flash(f'Restauración desde "{backup_filename}" completada.', 'success')
        flash('IMPORTANTE: Puede ser necesario reiniciar la aplicación para que todos los cambios surtan efecto.', 'warning')
            
    except Exception as e:
        flash(f'Error al restaurar: {e}', 'danger')

    return redirect(url_for('main.backup_restore'))

# app/routes.py

# ... (tus otras rutas) ...

@bp.route('/buscar_persona')
@login_required
def buscar_persona():
    """Endpoint de API para buscar personas por nombre."""
    search_term = request.args.get('q', '').strip()

    # Evitar búsquedas vacías o muy cortas para no sobrecargar la BD
    if len(search_term) < 3:
        return jsonify([])

    # Buscar personas cuyo nombre contenga el término de búsqueda (insensible a mayúsculas)
    # Usamos ilike para compatibilidad con PostgreSQL y otros dialectos.
    personas_encontradas = Persona.query.filter(
        Persona.nombre_persona.ilike(f"%{search_term}%")
    ).order_by(Persona.nombre_persona).limit(20).all()

    # Formatear los resultados para la respuesta JSON
    resultados = [
        {
            'id_persona': p.id_persona,
            'nombre_persona': p.nombre_persona,
            'dpto': p.dpto_ref.nombre_dpto
        }
        for p in personas_encontradas
    ]

    return jsonify(resultados)

# app/routes.py


# --- Ruta para la Importación Específica de Estudiantes desde Excel ---

@bp.route('/admin/importar_estudiantes_excel', methods=['GET', 'POST'])
@login_required
@admin_required
def importar_estudiantes_excel():
    if request.method == 'POST':
        if 'excel_file' not in request.files or not request.files['excel_file'].filename:
            flash('No se seleccionó ningún archivo.', 'warning')
            return redirect(request.url)
        
        file = request.files['excel_file']

        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Formato de archivo no válido. Por favor, sube un archivo de Excel (.xlsx).', 'danger')
            return redirect(request.url)

        try:
            updated_count = 0
            created_count = 0
            errors = []
            
            workbook = load_workbook(file)
            sheet = workbook.active

            tipos_persona_map = {tp.nombre_tipopersona: tp for tp in TipoPersona.query.all()}
            dptos_map = {d.nombre_dpto: d for d in Dpto.query.all()}
            tipos_control_map = {tc.nombre_control: tc for tc in TipoControl.query.all()}
            
            required_keys = ['Estudiante', 'Almuerzo Regular', 'Dieta Especial', 'No Aplica']
            if any(key not in (list(tipos_persona_map.keys()) + list(tipos_control_map.keys())) for key in required_keys if key != 'Estudiante'):
                 errors.append("Asegúrese de que los tipos de persona/control 'Estudiante', 'Almuerzo Regular', 'Dieta Especial' y 'No Aplica' existan en la base de datos.")

            if errors:
                return render_template('admin/importar_estudiantes_excel.html', title="Actualizar Estudiantes", errors=errors)

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue # Saltar filas completamente vacías
                seccion, nombre_raw, grupos, id_persona, sexo_raw = row[0:5]

                if not id_persona or not nombre_raw:
                    errors.append(f"Fila {i}: Faltan el ID o el Nombre del estudiante. Se omitirá esta fila.")
                    continue
                
                id_persona = str(id_persona).strip()
                
                try:
                    apellido, nombre = [part.strip() for part in nombre_raw.split(',', 1)]
                    nombre_formateado = f"{nombre} {apellido}"
                except (ValueError, TypeError):
                    nombre_formateado = str(nombre_raw).strip() if nombre_raw else ''
                
                sexo = 'M' if sexo_raw and 'MASCULINO' in str(sexo_raw).upper() else 'F'
                
                seccion_obj = dptos_map.get(str(seccion).strip()) if seccion else None
                if not seccion_obj:
                    errors.append(f"Fila {i} (ID: {id_persona}): El departamento '{seccion}' no existe o está vacío. Se omitirá esta fila.")
                    continue

                # --- LÓGICA DE BÚSQUEDA DE SUBCADENA (CONFIRMADA) ---
                grupos_str = str(grupos).upper() if grupos else ""
                if 'ALMUERZO NORMAL' in grupos_str:
                    control_obj = tipos_control_map.get('Almuerzo Regular')
                elif 'ALMUERZO ESPECIAL' in grupos_str:
                    control_obj = tipos_control_map.get('Dieta Especial')
                else:
                    control_obj = tipos_control_map.get('No Aplica')
                # --- FIN DE LA LÓGICA ---
                
                if not control_obj:
                    errors.append(f"Fila {i} (ID: {id_persona}): No se pudo asignar un tipo de control válido. Verifique que los tipos de control base existan.")
                    continue
                
                persona = Persona.query.get(id_persona)
                
                if persona:
                    persona.nombre_persona = nombre_formateado
                    persona.sexo = sexo
                    persona.dpto_id = seccion_obj.id_dpto
                    persona.control_id = control_obj.id_control
                    persona.tipo_persona_id = tipos_persona_map['Estudiante'].id_tipopersona
                    updated_count += 1
                else:
                    new_person = Persona(
                        id_persona=id_persona,
                        nombre_persona=nombre_formateado,
                        sexo=sexo,
                        dpto_id=seccion_obj.id_dpto,
                        control_id = control_obj.id_control,
                        tipo_persona_id = tipos_persona_map['Estudiante'].id_tipopersona
                    )
                    db.session.add(new_person)
                    created_count += 1
            
            if errors:
                db.session.rollback()
                flash("La importación se canceló debido a errores. Revisa la lista de problemas.", 'danger')
                return render_template('admin/importar_estudiantes_excel.html', title="Actualizar Estudiantes", errors=errors)
            
            db.session.commit()
            success_message = []
            if updated_count > 0: success_message.append(f"{updated_count} estudiantes actualizados")
            if created_count > 0: success_message.append(f"{created_count} estudiantes nuevos creados")
            
            if not success_message:
                flash('El archivo no contenía estudiantes para procesar.', 'warning')
            else:
                flash(f"Proceso completado: {', '.join(success_message)}.", 'success')

            return redirect(url_for('main.list_personas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Ocurrió un error inesperado al procesar el archivo: {e}', 'danger')
            return redirect(request.url)

    return render_template('admin/importar_estudiantes_excel.html', title="Actualizar Estudiantes")

# app/routes.py

# ... (tus otras rutas) ...

@bp.route('/usuarios/change_password', methods=['POST'])
@login_required
@admin_required
def change_password():
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    if not all([user_id, new_password, confirm_password]):
        flash('Todos los campos son requeridos.', 'danger')
        return redirect(url_for('main.list_usuarios'))
    
    if len(new_password) < 6:
        flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
        return redirect(url_for('main.list_usuarios'))
        
    if new_password != confirm_password:
        flash('Las contraseñas no coinciden.', 'danger')
        return redirect(url_for('main.list_usuarios'))
        
    user = Usuario.query.get(user_id)
    if not user:
        flash('Usuario no encontrado.', 'danger')
        return redirect(url_for('main.list_usuarios'))
    
    # Hashear y guardar la nueva contraseña
    user.password = bcrypt.generate_password_hash(new_password).decode('utf-8')
    db.session.commit()
    
    flash(f'La contraseña para el usuario "{user.username}" ha sido cambiada exitosamente.', 'success')
    return redirect(url_for('main.list_usuarios'))